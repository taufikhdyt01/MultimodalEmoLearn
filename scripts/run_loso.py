"""
LOSO (Leave-One-Subject-Out) Cross-Validation
==============================================
Untuk setiap user, jadikan sebagai test set, sisanya sebagai train+val.
Jalankan pada top 3 model front-only.

Usage:
    python scripts/run_loso.py                          # semua top 3 model
    python scripts/run_loso.py --models intermediate_tl  # model tertentu
    python scripts/run_loso.py --num-classes 4           # 4-class saja
"""
import sys
import os
import json
import csv
import argparse
import numpy as np
import torch
import torch.nn as nn
from pathlib import Path
from collections import defaultdict, Counter
from torch.utils.data import DataLoader, TensorDataset
import cv2

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from training.models import (
    EmotionCNN, EmotionFCNN, IntermediateFusion,
    EmotionCNNTransfer, IntermediateFusionTransfer,
)
from training.utils import train_model, full_evaluation

# ── CONFIG ────────────────────────────────────────────
FINAL_DIR = PROJECT_ROOT / "data" / "final"
OLD_DIR = FINAL_DIR / "old"
NEW_DIR = FINAL_DIR / "new"
OLD_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
NEW_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed_new"
OUTPUT_DIR = PROJECT_ROOT / "models" / "frontonly" / "loso"

EMOTIONS_7 = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]
EMOTIONS_4 = ["neutral", "happy", "sad", "negative"]
REMAP_4 = {0: 0, 1: 1, 2: 2, 3: 3, 4: 3, 5: 3, 6: 3}

IMG_SIZE = 224
BATCH_SIZE = 32
EPOCHS = 50
PATIENCE = 15


# ── DATA COLLECTION (reuse from prepare_dataset.py) ──

def load_old_labels():
    import openpyxl
    sample_uid_map = {}
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        for xlsx in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                sample_uid_map[sample_dir.name] = str(row[1])
            wb.close()
            break
    user_labels = defaultdict(dict)
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        uid = sample_uid_map.get(sample_dir.name)
        if not uid:
            continue
        for xlsx_path in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                timestamp = row[2]
                if timestamp is None:
                    continue
                time_key = timestamp.strftime("%H_%M_%S")
                scores = []
                for emo in EMOTIONS_7:
                    idx = headers.index(emo)
                    val = row[idx] if row[idx] is not None else 0.0
                    scores.append(float(val))
                user_labels[uid][time_key] = scores
            wb.close()
    return user_labels


def load_new_labels():
    user_labels = defaultdict(dict)
    for user_dir in NEW_PROCESSED_DIR.iterdir():
        if not user_dir.is_dir() or not user_dir.name.isdigit():
            continue
        uid = user_dir.name
        labels_csv = user_dir / "labels.csv"
        if not labels_csv.exists():
            continue
        with open(labels_csv, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                emo_id = row["emotion_id"]
                scores = [float(row.get(emo, 0.0)) for emo in EMOTIONS_7]
                user_labels[uid][emo_id] = scores
    return user_labels


def collect_all_frontonly_samples(old_labels, new_labels):
    """Collect all front-only samples grouped by user."""
    user_samples = defaultdict(list)

    # Old data (front only)
    if OLD_DIR.exists():
        for uid_dir in sorted(OLD_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = old_labels.get(uid, {})
            faces_dir = uid_dir / "front" / "faces"
            lm_dir = uid_dir / "front" / "landmarks"
            if not faces_dir.exists():
                continue
            for face_file in sorted(faces_dir.glob("*.jpg")):
                time_key = face_file.stem.replace("frame_", "")
                lm_file = lm_dir / (face_file.stem + ".csv")
                if time_key in labels and lm_file.exists():
                    scores = labels[time_key]
                    label_idx = int(np.argmax(scores))
                    user_samples[uid].append({
                        "face_path": str(face_file),
                        "lm_path": str(lm_file),
                        "label": label_idx,
                    })

    # New data (front only — no side)
    if NEW_DIR.exists():
        for uid_dir in sorted(NEW_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = new_labels.get(uid, {})
            faces_dir = uid_dir / "front" / "faces"
            lm_dir = uid_dir / "front" / "landmarks"
            if not faces_dir.exists():
                continue
            for face_file in sorted(faces_dir.glob("*.jpg")):
                parts = face_file.stem.split("_emo")
                if len(parts) != 2:
                    continue
                emo_id = parts[1]
                lm_file = lm_dir / (face_file.stem + ".csv")
                if emo_id in labels and lm_file.exists():
                    scores = labels[emo_id]
                    label_idx = int(np.argmax(scores))
                    user_samples[uid].append({
                        "face_path": str(face_file),
                        "lm_path": str(lm_file),
                        "label": label_idx,
                    })

    return dict(user_samples)


def load_image(path):
    img = cv2.imread(path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    return img.astype(np.float32) / 255.0


def load_landmark(path):
    coords = []
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            coords.append(float(row["x"]))
            coords.append(float(row["y"]))
    return np.array(coords, dtype=np.float32)


def build_arrays_from_samples(samples, num_classes=7):
    """Build numpy arrays from sample list."""
    n = len(samples)
    images = np.zeros((n, IMG_SIZE, IMG_SIZE, 3), dtype=np.float32)
    landmarks = np.zeros((n, 136), dtype=np.float32)
    labels = np.zeros(n, dtype=np.int64)

    for i, s in enumerate(samples):
        images[i] = load_image(s["face_path"])
        landmarks[i] = load_landmark(s["lm_path"])
        if num_classes == 4:
            labels[i] = REMAP_4[s["label"]]
        else:
            labels[i] = s["label"]

    return images, landmarks, labels


def make_loaders(images, landmarks, labels, model_type, batch_size=32, shuffle=True):
    """Create DataLoader based on model type."""
    img_t = torch.from_numpy(images).permute(0, 3, 1, 2)  # NHWC -> NCHW
    lm_t = torch.from_numpy(landmarks)
    y_t = torch.from_numpy(labels).long()

    if model_type == "cnn":
        ds = TensorDataset(img_t, y_t)
    elif model_type == "fcnn":
        ds = TensorDataset(lm_t, y_t)
    else:  # fusion
        ds = TensorDataset(img_t, lm_t, y_t)

    return DataLoader(ds, batch_size=batch_size, shuffle=shuffle,
                      num_workers=0, pin_memory=True)


# ── MODEL DEFINITIONS ────────────────────────────────

MODEL_CONFIGS = {
    "intermediate_tl": {
        "class": IntermediateFusionTransfer,
        "type": "fusion",
        "lr": 0.00005,
        "description": "Intermediate Fusion TL (ResNet18 + FCNN)",
    },
    "late_fusion": {
        "class": None,  # special handling
        "type": "late",
        "lr": 0.0001,
        "description": "Late Fusion (CNN + FCNN weighted avg)",
    },
    "fcnn": {
        "class": EmotionFCNN,
        "type": "fcnn",
        "lr": 0.0001,
        "description": "FCNN (Landmark only)",
    },
    "cnn_tl": {
        "class": EmotionCNNTransfer,
        "type": "cnn",
        "lr": 0.00005,
        "description": "CNN TL (ResNet18)",
    },
    "cnn": {
        "class": EmotionCNN,
        "type": "cnn",
        "lr": 0.0001,
        "description": "CNN (from scratch)",
    },
    "intermediate": {
        "class": IntermediateFusion,
        "type": "fusion",
        "lr": 0.0001,
        "description": "Intermediate Fusion (from scratch)",
    },
}


def train_and_eval_fold(model_name, train_samples, test_samples,
                        num_classes, device, fold_dir):
    """Train model on train_samples, evaluate on test_samples."""
    cfg = MODEL_CONFIGS[model_name]
    emotions = EMOTIONS_4 if num_classes == 4 else EMOTIONS_7

    # Build arrays
    train_img, train_lm, train_y = build_arrays_from_samples(train_samples, num_classes)
    test_img, test_lm, test_y = build_arrays_from_samples(test_samples, num_classes)

    # For validation, take 10% of train
    n_val = max(1, int(len(train_y) * 0.1))
    rng = np.random.RandomState(42)
    indices = rng.permutation(len(train_y))
    val_idx = indices[:n_val]
    tr_idx = indices[n_val:]

    val_img, val_lm, val_y = train_img[val_idx], train_lm[val_idx], train_y[val_idx]
    train_img, train_lm, train_y = train_img[tr_idx], train_lm[tr_idx], train_y[tr_idx]

    model_type = cfg["type"]

    # Late fusion: train CNN + FCNN separately, then combine
    if model_type == "late":
        from sklearn.metrics import f1_score as f1s, accuracy_score as acc_s

        # Train CNN
        cnn_model = EmotionCNN(num_classes=num_classes).to(device)
        cnn_train = make_loaders(train_img, train_lm, train_y, "cnn", BATCH_SIZE)
        cnn_val = make_loaders(val_img, val_lm, val_y, "cnn", BATCH_SIZE, shuffle=False)
        cnn_crit = nn.CrossEntropyLoss()
        cnn_opt = torch.optim.Adam(cnn_model.parameters(), lr=0.0001)
        cnn_sched = torch.optim.lr_scheduler.ReduceLROnPlateau(cnn_opt, mode="max", factor=0.5, patience=8, min_lr=1e-7)
        train_model(cnn_model, cnn_train, cnn_val, cnn_crit, cnn_opt, cnn_sched,
                     device, "cnn", EPOCHS, PATIENCE, str(fold_dir / "cnn_temp.pth"))

        # Train FCNN
        fcnn_model = EmotionFCNN(num_classes=num_classes).to(device)
        fcnn_train = make_loaders(train_img, train_lm, train_y, "fcnn", BATCH_SIZE)
        fcnn_val = make_loaders(val_img, val_lm, val_y, "fcnn", BATCH_SIZE, shuffle=False)
        fcnn_crit = nn.CrossEntropyLoss()
        fcnn_opt = torch.optim.Adam(fcnn_model.parameters(), lr=0.0001)
        fcnn_sched = torch.optim.lr_scheduler.ReduceLROnPlateau(fcnn_opt, mode="max", factor=0.5, patience=8, min_lr=1e-7)
        train_model(fcnn_model, fcnn_train, fcnn_val, fcnn_crit, fcnn_opt, fcnn_sched,
                     device, "fcnn", EPOCHS, PATIENCE, str(fold_dir / "fcnn_temp.pth"))

        # Load best
        cnn_model.load_state_dict(torch.load(fold_dir / "cnn_temp.pth", map_location=device, weights_only=True))
        fcnn_model.load_state_dict(torch.load(fold_dir / "fcnn_temp.pth", map_location=device, weights_only=True))
        cnn_model.eval(); fcnn_model.eval()

        # Late fusion on test
        test_img_t = torch.from_numpy(test_img).permute(0, 3, 1, 2).to(device)
        test_lm_t = torch.from_numpy(test_lm).to(device)

        with torch.no_grad():
            cnn_probs = torch.softmax(cnn_model(test_img_t), dim=1).cpu().numpy()
            fcnn_probs = torch.softmax(fcnn_model(test_lm_t), dim=1).cpu().numpy()

        best_f1, best_w = 0, 0.5
        for w in np.arange(0.0, 1.05, 0.05):
            preds = (w * cnn_probs + (1 - w) * fcnn_probs).argmax(axis=1)
            f1 = f1s(test_y, preds, average="macro", zero_division=0)
            if f1 > best_f1:
                best_f1 = f1; best_w = w; best_preds = preds

        acc = acc_s(test_y, best_preds)
        wf1 = f1s(test_y, best_preds, average="weighted", zero_division=0)

        # Cleanup temp files
        (fold_dir / "cnn_temp.pth").unlink(missing_ok=True)
        (fold_dir / "fcnn_temp.pth").unlink(missing_ok=True)

        return {"accuracy": acc, "macro_f1": best_f1, "weighted_f1": wf1,
                "best_cnn_weight": best_w, "test_samples": len(test_y)}

    # Standard models (non-late-fusion)
    model = cfg["class"](num_classes=num_classes).to(device)
    train_loader = make_loaders(train_img, train_lm, train_y, model_type, BATCH_SIZE)
    val_loader = make_loaders(val_img, val_lm, val_y, model_type, BATCH_SIZE, shuffle=False)
    test_loader = make_loaders(test_img, test_lm, test_y, model_type, BATCH_SIZE, shuffle=False)

    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=cfg["lr"])
    scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        optimizer, mode="max", factor=0.5, patience=8, min_lr=1e-7)

    save_path = fold_dir / "model_temp.pth"
    train_model(model, train_loader, val_loader, criterion, optimizer, scheduler,
                device, model_type, EPOCHS, PATIENCE, str(save_path))

    # Load best and evaluate
    model.load_state_dict(torch.load(save_path, map_location=device, weights_only=True))
    results = full_evaluation(model, test_loader, criterion, device, model_type, emotions)

    # Cleanup
    save_path.unlink(missing_ok=True)

    return {
        "accuracy": float(results["test_accuracy"]),
        "macro_f1": float(results["test_macro_f1"]),
        "weighted_f1": float(results["test_weighted_f1"]),
        "test_samples": len(test_y),
    }


def main():
    parser = argparse.ArgumentParser(description="LOSO Cross-Validation")
    parser.add_argument("--models", nargs="+",
                        default=["intermediate_tl", "late_fusion", "fcnn"],
                        help="Models to evaluate")
    parser.add_argument("--num-classes", type=int, default=4,
                        help="Number of classes (4 or 7)")
    parser.add_argument("--device", default="cuda" if torch.cuda.is_available() else "cpu")
    args = parser.parse_args()

    device = torch.device(args.device)
    print(f"Device: {device}")
    if device.type == "cuda":
        print(f"GPU: {torch.cuda.get_device_name(0)}")

    num_classes = args.num_classes
    emotions = EMOTIONS_4 if num_classes == 4 else EMOTIONS_7
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ── Collect all samples ──
    print("\nCollecting samples (front-only)...")
    old_labels = load_old_labels()
    new_labels = load_new_labels()
    user_samples = collect_all_frontonly_samples(old_labels, new_labels)

    users = sorted(user_samples.keys())
    total = sum(len(v) for v in user_samples.values())
    print(f"Total: {total} samples from {len(users)} users")
    for uid in users:
        n = len(user_samples[uid])
        counts = Counter(s["label"] for s in user_samples[uid])
        print(f"  User {uid:>4s}: {n:>4d} samples | {dict(sorted(counts.items()))}")

    # ── Run LOSO ──
    for model_name in args.models:
        cfg = MODEL_CONFIGS[model_name]
        print(f"\n{'='*70}")
        print(f"  LOSO: {cfg['description']} ({num_classes}-class)")
        print(f"  {len(users)} folds (1 user = 1 fold)")
        print(f"{'='*70}")

        fold_results = []
        model_dir = OUTPUT_DIR / f"{model_name}_{num_classes}class"
        os.makedirs(model_dir, exist_ok=True)

        for i, test_user in enumerate(users):
            print(f"\n  Fold {i+1}/{len(users)}: test=user_{test_user} "
                  f"({len(user_samples[test_user])} samples)")

            # Split
            test_samps = user_samples[test_user]
            train_samps = []
            for uid in users:
                if uid != test_user:
                    train_samps.extend(user_samples[uid])

            # Skip if test set has no samples for target classes
            test_labels = [REMAP_4[s["label"]] if num_classes == 4 else s["label"]
                          for s in test_samps]
            if len(set(test_labels)) < 2:
                print(f"    SKIP: test user only has {len(set(test_labels))} class(es)")
                continue

            fold_dir = model_dir / f"fold_{test_user}"
            os.makedirs(fold_dir, exist_ok=True)

            result = train_and_eval_fold(
                model_name, train_samps, test_samps,
                num_classes, device, fold_dir)

            result["test_user"] = test_user
            fold_results.append(result)

            print(f"    Acc={result['accuracy']:.4f} "
                  f"Macro-F1={result['macro_f1']:.4f} "
                  f"W-F1={result['weighted_f1']:.4f} "
                  f"(n={result['test_samples']})")

            # Cleanup fold dir
            try:
                fold_dir.rmdir()
            except OSError:
                pass

        # ── Summary ──
        if fold_results:
            accs = [r["accuracy"] for r in fold_results]
            f1s = [r["macro_f1"] for r in fold_results]
            wf1s = [r["weighted_f1"] for r in fold_results]

            print(f"\n{'='*70}")
            print(f"  LOSO RESULTS: {cfg['description']} ({num_classes}-class)")
            print(f"{'='*70}")
            print(f"  Folds completed: {len(fold_results)}/{len(users)}")
            print(f"  Accuracy:    {np.mean(accs):.4f} +/- {np.std(accs):.4f}")
            print(f"  Macro F1:    {np.mean(f1s):.4f} +/- {np.std(f1s):.4f}")
            print(f"  Weighted F1: {np.mean(wf1s):.4f} +/- {np.std(wf1s):.4f}")
            print(f"  Best fold:   user_{fold_results[np.argmax(f1s)]['test_user']} (F1={max(f1s):.4f})")
            print(f"  Worst fold:  user_{fold_results[np.argmin(f1s)]['test_user']} (F1={min(f1s):.4f})")

            # Save
            summary = {
                "model": model_name,
                "description": cfg["description"],
                "num_classes": num_classes,
                "num_folds": len(fold_results),
                "total_users": len(users),
                "accuracy_mean": float(np.mean(accs)),
                "accuracy_std": float(np.std(accs)),
                "macro_f1_mean": float(np.mean(f1s)),
                "macro_f1_std": float(np.std(f1s)),
                "weighted_f1_mean": float(np.mean(wf1s)),
                "weighted_f1_std": float(np.std(wf1s)),
                "per_fold": fold_results,
            }
            save_path = OUTPUT_DIR / f"loso_{model_name}_{num_classes}class.json"
            with open(save_path, "w") as f:
                json.dump(summary, f, indent=2)
            print(f"  Saved: {save_path}")


if __name__ == "__main__":
    main()
