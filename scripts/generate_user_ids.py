"""
Generate user_ids npy files for existing dataset_frontonly.
Needed for LOSO cross-validation without raw data.

Reads from data/final/ (same as prepare_dataset.py) and saves
user_ids_{train,val,test}.npy alongside existing numpy arrays.

Usage:
    python scripts/generate_user_ids.py
"""
import sys
import csv
import json
import numpy as np
from pathlib import Path
from collections import defaultdict

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

FINAL_DIR = PROJECT_ROOT / "data" / "final"
OLD_DIR = FINAL_DIR / "old"
NEW_DIR = FINAL_DIR / "new"
OLD_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed"
NEW_PROCESSED_DIR = PROJECT_ROOT / "data" / "processed_new"
DATASET_DIR = PROJECT_ROOT / "data" / "dataset_frontonly"

EMOTIONS = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]


def load_old_labels():
    import openpyxl
    sample_uid_map = {}
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        for xlsx in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                sample_uid_map[sample_dir.name] = str(row[1])
            wb.close()
            break
    user_labels = defaultdict(dict)
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        uid = sample_uid_map.get(sample_dir.name)
        if not uid:
            continue
        for xlsx_path in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                timestamp = row[2]
                if timestamp is None:
                    continue
                time_key = timestamp.strftime("%H_%M_%S")
                scores = []
                for emo in EMOTIONS:
                    idx = headers.index(emo)
                    val = row[idx] if row[idx] is not None else 0.0
                    scores.append(float(val))
                user_labels[uid][time_key] = scores
            wb.close()
    return user_labels


def load_new_labels():
    user_labels = defaultdict(dict)
    for user_dir in NEW_PROCESSED_DIR.iterdir():
        if not user_dir.is_dir() or not user_dir.name.isdigit():
            continue
        uid = user_dir.name
        labels_csv = user_dir / "labels.csv"
        if not labels_csv.exists():
            continue
        with open(labels_csv, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                emo_id = row["emotion_id"]
                scores = [float(row.get(emo, 0.0)) for emo in EMOTIONS]
                user_labels[uid][emo_id] = scores
    return user_labels


def collect_samples_with_uids(old_labels, new_labels):
    """Collect front-only samples preserving order matching prepare_dataset.py."""
    samples = []  # list of (uid, face_path, lm_path, scores)

    # Old data (front only) — same order as prepare_dataset.py
    if OLD_DIR.exists():
        for uid_dir in sorted(OLD_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = old_labels.get(uid, {})
            faces_dir = uid_dir / "front" / "faces"
            lm_dir = uid_dir / "front" / "landmarks"
            if not faces_dir.exists():
                continue
            for face_file in sorted(faces_dir.glob("*.jpg")):
                time_key = face_file.stem.replace("frame_", "")
                lm_file = lm_dir / (face_file.stem + ".csv")
                if time_key in labels and lm_file.exists():
                    scores = labels[time_key]
                    samples.append((uid, str(face_file), str(lm_file), scores))

    # New data (front only — no side)
    if NEW_DIR.exists():
        for uid_dir in sorted(NEW_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = new_labels.get(uid, {})
            faces_dir = uid_dir / "front" / "faces"
            lm_dir = uid_dir / "front" / "landmarks"
            if not faces_dir.exists():
                continue
            for face_file in sorted(faces_dir.glob("*.jpg")):
                parts = face_file.stem.split("_emo")
                if len(parts) != 2:
                    continue
                emo_id = parts[1]
                lm_file = lm_dir / (face_file.stem + ".csv")
                if emo_id in labels and lm_file.exists():
                    scores = labels[emo_id]
                    samples.append((uid, str(face_file), str(lm_file), scores))

    return samples


def main():
    print("Loading labels...")
    old_labels = load_old_labels()
    new_labels = load_new_labels()

    print("Collecting samples (front-only, same order as prepare_dataset)...")
    all_samples = collect_samples_with_uids(old_labels, new_labels)
    print(f"Total: {len(all_samples)} samples")

    # Load dataset_info to get user split
    info = json.load(open(DATASET_DIR / "dataset_info.json"))
    train_users = set(info["train"]["users"])
    val_users = set(info["val"]["users"])
    test_users = set(info["test"]["users"])

    # Split samples by user (same logic as prepare_dataset)
    train_uids, val_uids, test_uids = [], [], []

    for uid, face_path, lm_path, scores in all_samples:
        if uid in train_users:
            train_uids.append(uid)
        elif uid in val_users:
            val_uids.append(uid)
        elif uid in test_users:
            test_uids.append(uid)

    # Verify counts match
    y_train = np.load(DATASET_DIR / "y_train.npy")
    y_val = np.load(DATASET_DIR / "y_val.npy")
    y_test = np.load(DATASET_DIR / "y_test.npy")

    print(f"\nVerification:")
    print(f"  Train: {len(train_uids)} uids vs {len(y_train)} labels {'OK' if len(train_uids)==len(y_train) else 'MISMATCH!'}")
    print(f"  Val:   {len(val_uids)} uids vs {len(y_val)} labels {'OK' if len(val_uids)==len(y_val) else 'MISMATCH!'}")
    print(f"  Test:  {len(test_uids)} uids vs {len(y_test)} labels {'OK' if len(test_uids)==len(y_test) else 'MISMATCH!'}")

    if len(train_uids) != len(y_train) or len(val_uids) != len(y_val) or len(test_uids) != len(y_test):
        print("ERROR: Count mismatch! Cannot save user_ids.")
        return

    # Save
    for split, uids in [("train", train_uids), ("val", val_uids), ("test", test_uids)]:
        path = DATASET_DIR / f"user_ids_{split}.npy"
        np.save(path, np.array(uids))
        print(f"  Saved {path} ({len(uids)} entries, {len(set(uids))} unique users)")

    # Also save a complete user mapping for LOSO
    all_uids = train_uids + val_uids + test_uids
    all_labels = np.concatenate([y_train, y_val, y_test])
    np.save(DATASET_DIR / "user_ids_all.npy", np.array(all_uids))
    np.save(DATASET_DIR / "y_all.npy", all_labels)
    print(f"\n  Saved user_ids_all.npy ({len(all_uids)} entries)")
    print(f"  Saved y_all.npy ({len(all_labels)} entries)")

    print("\nDone!")


if __name__ == "__main__":
    main()
