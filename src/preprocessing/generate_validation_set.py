"""
Generate Validation Set for Expert Review
===========================================
Ambil sampel representatif dari dataset untuk validasi ahli psikologi.

Strategi:
- Semua sample non-neutral (rare emotions penting untuk divalidasi)
- Random sample dari neutral (sebagai kontrol)
- Export ke Excel + folder gambar untuk review

Usage:
    python src/preprocessing/generate_validation_set.py
    python src/preprocessing/generate_validation_set.py --neutral-sample 400
"""

import os
import csv
import json
import shutil
import argparse
import numpy as np
from pathlib import Path
from collections import defaultdict, Counter

# ============== KONFIGURASI ==============
FINAL_DIR = Path("data/final")
OLD_DIR = FINAL_DIR / "old"
NEW_DIR = FINAL_DIR / "new"
OLD_PROCESSED_DIR = Path("data/processed")
NEW_PROCESSED_DIR = Path("data/processed_new")
OUTPUT_DIR = Path("data/validation")

EMOTIONS = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]
NEUTRAL_SAMPLE_SIZE = 400  # jumlah neutral yang di-sample
RANDOM_SEED = 42
# ==========================================


def load_old_labels():
    """Load emotion labels dari data lama."""
    import openpyxl

    sample_uid_map = {}
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        for xlsx in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                sample_uid_map[sample_dir.name] = str(row[1])
            wb.close()
            break

    user_labels = defaultdict(dict)
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        uid = sample_uid_map.get(sample_dir.name)
        if not uid:
            continue
        for xlsx_path in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                timestamp = row[2]
                if timestamp is None:
                    continue
                time_key = timestamp.strftime("%H_%M_%S")
                scores = []
                for emo in EMOTIONS:
                    idx = headers.index(emo)
                    val = row[idx] if row[idx] is not None else 0.0
                    scores.append(float(val))
                user_labels[uid][time_key] = scores
            wb.close()
    return user_labels


def load_new_labels():
    """Load emotion labels dari data baru."""
    user_labels = defaultdict(dict)
    for user_dir in NEW_PROCESSED_DIR.iterdir():
        if not user_dir.is_dir() or not user_dir.name.isdigit():
            continue
        uid = user_dir.name
        labels_csv = user_dir / "labels.csv"
        if not labels_csv.exists():
            continue
        with open(labels_csv, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                emo_id = row["emotion_id"]
                scores = [float(row.get(emo, 0.0)) for emo in EMOTIONS]
                user_labels[uid][emo_id] = scores
    return user_labels


def collect_all_samples(old_labels, new_labels):
    """Kumpulkan semua samples dengan metadata lengkap."""
    samples = []

    # Old data
    if OLD_DIR.exists():
        for uid_dir in sorted(OLD_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = old_labels.get(uid, {})
            faces_dir = uid_dir / "front" / "faces"
            if not faces_dir.exists():
                continue
            for face_file in sorted(faces_dir.glob("*.jpg")):
                time_key = face_file.stem.replace("frame_", "")
                if time_key in labels:
                    scores = labels[time_key]
                    dominant_idx = int(np.argmax(scores))
                    samples.append({
                        "user_id": uid,
                        "source": "old",
                        "angle": "front",
                        "face_path": str(face_file),
                        "filename": face_file.name,
                        "scores": scores,
                        "auto_label": EMOTIONS[dominant_idx],
                        "auto_confidence": scores[dominant_idx],
                    })

    # New data
    if NEW_DIR.exists():
        for uid_dir in sorted(NEW_DIR.iterdir()):
            if not uid_dir.is_dir():
                continue
            uid = uid_dir.name
            labels = new_labels.get(uid, {})
            for angle in ["front", "side"]:
                faces_dir = uid_dir / angle / "faces"
                if not faces_dir.exists():
                    continue
                for face_file in sorted(faces_dir.glob("*.jpg")):
                    parts = face_file.stem.split("_emo")
                    if len(parts) != 2:
                        continue
                    emo_id = parts[1]
                    if emo_id in labels:
                        scores = labels[emo_id]
                        dominant_idx = int(np.argmax(scores))
                        samples.append({
                            "user_id": uid,
                            "source": "new",
                            "angle": angle,
                            "face_path": str(face_file),
                            "filename": face_file.name,
                            "scores": scores,
                            "auto_label": EMOTIONS[dominant_idx],
                            "auto_confidence": scores[dominant_idx],
                        })

    return samples


def select_validation_samples(samples, neutral_sample_size, seed, strategy="all_rare",
                              target_pct=None, min_per_class=30):
    """Pilih samples untuk validasi.

    Strategies:
    - "all_rare": Semua non-neutral + sample neutral (default lama)
    - "stratified": Stratified sampling ~target_pct% dari total,
                    kelas kecil tetap semua, kelas besar di-sample,
                    minimal min_per_class per kelas.
    """
    rng = np.random.RandomState(seed)

    # Group by emotion
    by_emotion = defaultdict(list)
    for s in samples:
        by_emotion[s["auto_label"]].append(s)

    selected = []

    if strategy == "stratified" and target_pct is not None:
        total_target = int(len(samples) * target_pct / 100)

        for emo in EMOTIONS:
            emo_samples = by_emotion.get(emo, [])
            if len(emo_samples) == 0:
                continue

            # Kelas kecil (< min_per_class * 2): ambil semua
            if len(emo_samples) <= min_per_class * 2:
                selected.extend(emo_samples)
            else:
                # Proportional sample, tapi minimal min_per_class
                n_target = max(min_per_class, int(len(emo_samples) * target_pct / 100))
                n_target = min(n_target, len(emo_samples))
                indices = rng.choice(len(emo_samples), n_target, replace=False)
                selected.extend([emo_samples[i] for i in indices])
    else:
        # Strategy: all_rare (semua non-neutral + sample neutral)
        for emo in EMOTIONS:
            emo_samples = by_emotion.get(emo, [])
            if emo == "neutral":
                n_sample = min(neutral_sample_size, len(emo_samples))
                indices = rng.choice(len(emo_samples), n_sample, replace=False)
                selected.extend([emo_samples[i] for i in indices])
            else:
                selected.extend(emo_samples)

    rng.shuffle(selected)
    return selected


def export_to_excel(samples, output_path):
    """Export validation set ke Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation"

    # Headers
    headers = [
        "No", "Image File", "User ID", "Source", "Angle",
        "Auto Label", "Confidence",
        "Neutral", "Happy", "Sad", "Angry", "Fearful", "Disgusted", "Surprised",
        "Expert Label", "Expert Notes"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Color coding for emotions
    emotion_colors = {
        "neutral": "D9E2F3",
        "happy": "C6EFCE",
        "sad": "B4C7E7",
        "angry": "FFC7CE",
        "fearful": "E2BFCC",
        "disgusted": "D9D2E9",
        "surprised": "FFF2CC",
    }

    # Data rows
    for i, s in enumerate(samples, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=s["filename"]).border = thin_border
        ws.cell(row=row, column=3, value=s["user_id"]).border = thin_border
        ws.cell(row=row, column=4, value=s["source"]).border = thin_border
        ws.cell(row=row, column=5, value=s["angle"]).border = thin_border

        # Auto label with color
        label_cell = ws.cell(row=row, column=6, value=s["auto_label"])
        label_cell.border = thin_border
        color = emotion_colors.get(s["auto_label"], "FFFFFF")
        label_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        ws.cell(row=row, column=7, value=round(s["auto_confidence"], 6)).border = thin_border

        # Scores
        for j, score in enumerate(s["scores"]):
            ws.cell(row=row, column=8 + j, value=round(score, 8)).border = thin_border

        # Expert columns (kosong, diisi ahli)
        expert_cell = ws.cell(row=row, column=15, value="")
        expert_cell.border = thin_border
        expert_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

        notes_cell = ws.cell(row=row, column=16, value="")
        notes_cell.border = thin_border

    # Column widths
    widths = [5, 40, 8, 6, 6, 12, 12, 10, 10, 10, 10, 10, 10, 10, 15, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else None].width = width

    # Auto filter
    ws.auto_filter.ref = f"A1:P{len(samples) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Add instruction sheet
    ws_info = wb.create_sheet("Instructions")
    instructions = [
        "PETUNJUK VALIDASI EMOSI",
        "",
        "1. Buka folder 'images/' yang berisi cropped face untuk setiap sample",
        "2. Lihat kolom 'Auto Label' dan 'Confidence' untuk label otomatis",
        "3. Bandingkan dengan ekspresi wajah di gambar",
        "4. Isi kolom 'Expert Label' dengan salah satu:",
        "   - neutral, happy, sad, angry, fearful, disgusted, surprised",
        "   - Atau kosongkan jika setuju dengan Auto Label",
        "5. Tambahkan catatan di 'Expert Notes' jika perlu",
        "",
        f"Total sample untuk divalidasi: {len(samples)}",
        "",
        "Emotion Distribution dalam set validasi:",
    ]
    counts = Counter(s["auto_label"] for s in samples)
    for emo in EMOTIONS:
        instructions.append(f"  {emo}: {counts.get(emo, 0)}")

    for i, line in enumerate(instructions, 1):
        ws_info.cell(row=i, column=1, value=line)
    ws_info.column_dimensions["A"].width = 70

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate Validation Set")
    parser.add_argument("--strategy", choices=["all_rare", "stratified"], default="stratified",
                        help="Strategi sampling: all_rare (semua non-neutral + sample neutral) "
                             "atau stratified (proporsional per kelas)")
    parser.add_argument("--target-pct", type=float, default=10.0,
                        help="Target persentase dari total dataset (untuk stratified, default: 10)")
    parser.add_argument("--min-per-class", type=int, default=30,
                        help="Minimum sample per kelas emosi (default: 30)")
    parser.add_argument("--neutral-sample", type=int, default=NEUTRAL_SAMPLE_SIZE,
                        help="Jumlah neutral sample (untuk all_rare, default: 400)")
    parser.add_argument("--seed", type=int, default=RANDOM_SEED)
    parser.add_argument("--output", type=str, default=None,
                        help="Output directory (default: auto-named)")
    args = parser.parse_args()

    # Auto-name output dir
    if args.output:
        output_dir = Path(args.output)
    else:
        output_dir = Path(f"data/validation_{args.strategy}_{int(args.target_pct)}pct")
    os.makedirs(output_dir, exist_ok=True)
    images_dir = output_dir / "images"
    os.makedirs(images_dir, exist_ok=True)

    print("=" * 60)
    print("GENERATE VALIDATION SET")
    print(f"  Strategy: {args.strategy}, Target: {args.target_pct}%")
    print("=" * 60)

    # 1. Load labels
    print("\n[1/4] Loading labels...")
    old_labels = load_old_labels()
    new_labels = load_new_labels()

    # 2. Collect all samples
    print("[2/4] Collecting samples...")
    all_samples = collect_all_samples(old_labels, new_labels)
    print(f"  Total samples: {len(all_samples)}")

    # 3. Select validation samples
    print(f"[3/4] Selecting validation samples...")
    selected = select_validation_samples(
        all_samples, args.neutral_sample, args.seed,
        strategy=args.strategy, target_pct=args.target_pct,
        min_per_class=args.min_per_class
    )

    counts = Counter(s["auto_label"] for s in selected)
    print(f"  Selected: {len(selected)} samples ({len(selected)/len(all_samples)*100:.1f}% of total)")
    for emo in EMOTIONS:
        total_emo = sum(1 for s in all_samples if s["auto_label"] == emo)
        count = counts.get(emo, 0)
        pct = count / total_emo * 100 if total_emo > 0 else 0
        print(f"    {emo:>10s}: {count:>4d} / {total_emo:>5d} ({pct:.0f}%)")

    # 4. Export
    print(f"[4/4] Exporting...")

    for i, s in enumerate(selected):
        dest = images_dir / f"{i+1:04d}_{s['auto_label']}_{s['filename']}"
        if not dest.exists():
            shutil.copy2(s["face_path"], dest)

    excel_path = output_dir / "validation_sheet.xlsx"
    export_to_excel(selected, excel_path)

    summary = {
        "strategy": args.strategy,
        "target_pct": args.target_pct,
        "min_per_class": args.min_per_class,
        "total_samples": len(selected),
        "seed": args.seed,
        "distribution": dict(counts),
        "total_dataset_size": len(all_samples),
        "sampling_ratio": round(len(selected) / len(all_samples) * 100, 1),
    }
    with open(output_dir / "validation_info.json", "w") as f:
        json.dump(summary, f, indent=2)

    print(f"\n{'=' * 60}")
    print(f"SELESAI!")
    print(f"  Total: {len(selected)} samples ({summary['sampling_ratio']}% of dataset)")
    print(f"  Excel: {excel_path}")
    print(f"  Images: {images_dir}/")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
