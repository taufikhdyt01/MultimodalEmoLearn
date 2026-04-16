"""
Prepare Dataset for Training
==============================
Gabungkan data lama (20 user) + baru (17 user), match dengan emotion labels,
split train/val/test stratified by user, simpan sebagai numpy arrays.

Output:
    data/dataset/
    ├── X_train_images.npy     # (N, 224, 224, 3) float32 [0-1]
    ├── X_train_landmarks.npy  # (N, 136) float32
    ├── y_train.npy            # (N,) int
    ├── X_val_images.npy
    ├── X_val_landmarks.npy
    ├── y_val.npy
    ├── X_test_images.npy
    ├── X_test_landmarks.npy
    ├── y_test.npy
    ├── label_map.json         # {0: "neutral", 1: "happy", ...}
    └── dataset_info.json      # statistik dataset

Usage:
    python src/preprocessing/prepare_dataset.py
    python src/preprocessing/prepare_dataset.py --include-side
"""

import os
import csv
import json
import argparse
import numpy as np
import cv2
from pathlib import Path
from collections import defaultdict, Counter

# ============== KONFIGURASI ==============
FINAL_DIR = Path("data/final")
OLD_DIR = FINAL_DIR / "old"
NEW_DIR = FINAL_DIR / "new"
OUTPUT_DIR = Path("data/dataset")

# Data lama: processed folder untuk ambil emotion labels
OLD_PROCESSED_DIR = Path("data/processed")
# Data baru: labels dari batch_video_processor
NEW_PROCESSED_DIR = Path("data/processed_new")

EMOTIONS = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]
LABEL_MAP = {emo: i for i, emo in enumerate(EMOTIONS)}

IMG_SIZE = 224
SPLIT_RATIO = (0.8, 0.1, 0.1)  # train, val, test
RANDOM_SEED = 42
MIN_CONFIDENCE = 0.0  # default: no filtering. Set via --min-confidence
# ==========================================


def get_dominant_emotion(scores):
    """Tentukan emosi dominan dari 7 skor."""
    idx = np.argmax(scores)
    return idx, scores[idx]


def load_old_labels():
    """Load emotion labels dari data lama (cleaned_data.xlsx per sample).
    Returns: dict { user_id: { "HH_MM_SS": emotion_scores, ... } }
    """
    import openpyxl

    # Mapping Sample -> user_id
    sample_uid_map = {}
    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        for xlsx in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                sample_uid_map[sample_dir.name] = str(row[1])
            wb.close()
            break

    # Load all labels grouped by user_id + timestamp
    user_labels = defaultdict(dict)

    for sample_dir in OLD_PROCESSED_DIR.iterdir():
        if not sample_dir.is_dir() or not sample_dir.name.startswith("Sample"):
            continue
        uid = sample_uid_map.get(sample_dir.name)
        if not uid:
            continue

        for xlsx_path in sample_dir.rglob("cleaned_data.xlsx"):
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

            for row in ws.iter_rows(min_row=2, values_only=True):
                timestamp = row[2]  # datetime object
                if timestamp is None:
                    continue
                # Frame name format: frame_HH_MM_SS.jpg
                time_key = timestamp.strftime("%H_%M_%S")

                scores = []
                for emo in EMOTIONS:
                    idx = headers.index(emo)
                    val = row[idx] if row[idx] is not None else 0.0
                    scores.append(float(val))

                user_labels[uid][time_key] = np.array(scores, dtype=np.float32)
            wb.close()

    return user_labels


def load_new_labels():
    """Load emotion labels dari data baru (labels.csv per user).
    Returns: dict { user_id: { emotion_id: emotion_scores, ... } }
    """
    user_labels = defaultdict(dict)

    for user_dir in NEW_PROCESSED_DIR.iterdir():
        if not user_dir.is_dir() or not user_dir.name.isdigit():
            continue
        uid = user_dir.name
        labels_csv = user_dir / "labels.csv"
        if not labels_csv.exists():
            continue

        with open(labels_csv, "r") as f:
            reader = csv.DictReader(f)
            for row in reader:
                emo_id = row["emotion_id"]
                scores = []
                for emo in EMOTIONS:
                    scores.append(float(row.get(emo, 0.0)))
                user_labels[uid][emo_id] = np.array(scores, dtype=np.float32)

    return user_labels


def collect_old_samples(user_labels, min_confidence=0.0):
    """Kumpulkan samples dari data lama.
    Returns: list of (user_id, face_path, landmark_path, emotion_scores)
    """
    samples = []
    matched = 0
    unmatched = 0
    filtered = 0

    if not OLD_DIR.exists():
        return samples

    for uid_dir in sorted(OLD_DIR.iterdir()):
        if not uid_dir.is_dir():
            continue
        uid = uid_dir.name
        labels = user_labels.get(uid, {})

        faces_dir = uid_dir / "front" / "faces"
        lm_dir = uid_dir / "front" / "landmarks"
        if not faces_dir.exists():
            continue

        for face_file in sorted(faces_dir.glob("*.jpg")):
            stem = face_file.stem  # frame_HH_MM_SS
            time_key = stem.replace("frame_", "")  # HH_MM_SS
            lm_file = lm_dir / f"{stem}.csv"

            if time_key in labels and lm_file.exists():
                scores = labels[time_key]
                confidence = float(np.max(scores))
                if confidence < min_confidence:
                    filtered += 1
                    continue
                samples.append((uid, str(face_file), str(lm_file), scores))
                matched += 1
            else:
                unmatched += 1

    print(f"  Old data: {matched} matched, {unmatched} unmatched, {filtered} filtered (conf < {min_confidence})")
    return samples


def collect_new_samples(user_labels, include_side=False, min_confidence=0.0):
    """Kumpulkan samples dari data baru.
    Returns: list of (user_id, face_path, landmark_path, emotion_scores)
    """
    samples = []
    matched = 0
    unmatched = 0
    filtered = 0

    if not NEW_DIR.exists():
        return samples

    angles = ["front", "side"] if include_side else ["front"]

    for uid_dir in sorted(NEW_DIR.iterdir()):
        if not uid_dir.is_dir():
            continue
        uid = uid_dir.name
        labels = user_labels.get(uid, {})

        for angle in angles:
            faces_dir = uid_dir / angle / "faces"
            lm_dir = uid_dir / angle / "landmarks"
            if not faces_dir.exists():
                continue

            for face_file in sorted(faces_dir.glob("*.jpg")):
                stem = face_file.stem  # frame_YYYYMMDD_HHMMSS_emoXXXXX
                lm_file = lm_dir / f"{stem}.csv"

                # Extract emotion_id from filename
                parts = stem.split("_emo")
                if len(parts) == 2:
                    emo_id = parts[1]
                else:
                    unmatched += 1
                    continue

                if emo_id in labels and lm_file.exists():
                    scores = labels[emo_id]
                    confidence = float(np.max(scores))
                    if confidence < min_confidence:
                        filtered += 1
                        continue
                    samples.append((uid, str(face_file), str(lm_file), scores))
                    matched += 1
                else:
                    unmatched += 1

    print(f"  New data: {matched} matched, {unmatched} unmatched, {filtered} filtered (conf < {min_confidence})")
    return samples


def load_image(path):
    """Load dan normalize gambar ke float32 [0-1]."""
    img = cv2.imread(path)
    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    return img.astype(np.float32) / 255.0


def load_landmark(path):
    """Load 68 landmark dari CSV, return flat array (136,)."""
    coords = []
    with open(path, "r") as f:
        reader = csv.DictReader(f)
        for row in reader:
            coords.append(float(row["x"]))
            coords.append(float(row["y"]))
    return np.array(coords, dtype=np.float32)


def split_by_user(samples, split_ratio, seed):
    """Split samples berdasarkan user_id dengan strategi:
    1. Hitung emosi per user
    2. Coba banyak seed, pilih yang semua 7 emosi ada di semua split
    3. Jika tidak ditemukan, gunakan best-effort

    Ini mencegah data leaking DAN memastikan emosi langka ada di semua split.
    """
    emotions_list = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]
    rare_emotions = {"fearful", "disgusted"}

    # Group by user
    user_samples = defaultdict(list)
    for s in samples:
        user_samples[s[0]].append(s)

    # Hitung emosi per user
    user_emotion_counts = {}
    for uid, samps in user_samples.items():
        counts = Counter()
        for _, _, _, scores in samps:
            counts[emotions_list[np.argmax(scores)]] += 1
        user_emotion_counts[uid] = counts

    # Identifikasi rare vs normal users
    rare_users = sorted([uid for uid, c in user_emotion_counts.items()
                         if any(c.get(e, 0) > 0 for e in rare_emotions)])
    normal_users = sorted([uid for uid in user_samples if uid not in rare_users])

    def try_split(s):
        rng = np.random.RandomState(s)
        r, n = list(rare_users), list(normal_users)
        rng.shuffle(r); rng.shuffle(n)

        nr = len(r); nn = len(n)
        nr_train = max(1, int(nr * split_ratio[0]))
        nr_val = max(1, int(nr * split_ratio[1]))
        nn_train = int(nn * split_ratio[0])
        nn_val = int(nn * split_ratio[1])

        t_u = set(r[:nr_train]) | set(n[:nn_train])
        v_u = set(r[nr_train:nr_train + nr_val]) | set(n[nn_train:nn_train + nn_val])
        e_u = set(r[nr_train + nr_val:]) | set(n[nn_train + nn_val:])
        return t_u, v_u, e_u

    def check_split(t_u, v_u, e_u):
        min_count = float('inf')
        for su in [t_u, v_u, e_u]:
            sc = Counter()
            for u in su:
                for e, c in user_emotion_counts[u].items():
                    sc[e] += c
            for e in emotions_list:
                min_count = min(min_count, sc.get(e, 0))
        return min_count

    # Cari seed terbaik
    best_seed, best_min = seed, -1
    for s in range(2000):
        t_u, v_u, e_u = try_split(s)
        mc = check_split(t_u, v_u, e_u)
        if mc > best_min:
            best_min = mc
            best_seed = s
        if mc >= 1:  # semua emosi ada di semua split
            break

    train_users, val_users, test_users = try_split(best_seed)
    print(f"  Split seed used: {best_seed} (min count per emotion per split: {best_min})")
    print(f"  Rare-emotion users ({len(rare_users)}): "
          f"train={len(set(rare_users) & train_users)}, "
          f"val={len(set(rare_users) & val_users)}, "
          f"test={len(set(rare_users) & test_users)}")

    train, val, test = [], [], []
    for uid in user_samples:
        if uid in train_users:
            train.extend(user_samples[uid])
        elif uid in val_users:
            val.extend(user_samples[uid])
        else:
            test.extend(user_samples[uid])

    return train, val, test, train_users, val_users, test_users


def build_arrays(samples, desc=""):
    """Convert list of samples ke numpy arrays."""
    n = len(samples)
    if n == 0:
        return np.array([]), np.array([]), np.array([])

    images = np.zeros((n, IMG_SIZE, IMG_SIZE, 3), dtype=np.float32)
    landmarks = np.zeros((n, 136), dtype=np.float32)
    labels = np.zeros(n, dtype=np.int32)

    for i, (uid, face_path, lm_path, scores) in enumerate(samples):
        images[i] = load_image(face_path)
        landmarks[i] = load_landmark(lm_path)
        labels[i] = get_dominant_emotion(scores)[0]

        if (i + 1) % 500 == 0 or (i + 1) == n:
            print(f"    {desc}: {i + 1}/{n}")

    return images, landmarks, labels


def main():
    parser = argparse.ArgumentParser(description="Prepare Dataset for Training")
    parser.add_argument("--include-side", action="store_true",
                        help="Sertakan data side view dari user baru")
    parser.add_argument("--output", type=str, default=str(OUTPUT_DIR))
    parser.add_argument("--seed", type=int, default=RANDOM_SEED)
    parser.add_argument("--min-confidence", type=float, default=0.0,
                        help="Minimum confidence score (0-1). Samples below this are excluded.")
    args = parser.parse_args()

    output = Path(args.output)
    os.makedirs(output, exist_ok=True)

    print("=" * 60)
    print("PREPARE DATASET")
    print("=" * 60)

    # 1. Load emotion labels
    print("\n[1/5] Loading emotion labels...")
    old_labels = load_old_labels()
    new_labels = load_new_labels()
    print(f"  Old: {len(old_labels)} users, "
          f"New: {len(new_labels)} users")

    # 2. Collect samples
    min_conf = args.min_confidence
    print(f"\n[2/5] Collecting samples... (min_confidence={min_conf})")
    old_samples = collect_old_samples(old_labels, min_confidence=min_conf)
    new_samples = collect_new_samples(new_labels, include_side=args.include_side, min_confidence=min_conf)
    all_samples = old_samples + new_samples
    print(f"  Total: {len(all_samples)} samples "
          f"({len(old_samples)} old + {len(new_samples)} new)")

    if not all_samples:
        print("ERROR: Tidak ada samples ditemukan!")
        return

    # 3. Emotion distribution
    print("\n[3/5] Emotion distribution:")
    emotion_counts = Counter()
    for _, _, _, scores in all_samples:
        emo_idx = np.argmax(scores)
        emotion_counts[EMOTIONS[emo_idx]] += 1

    for emo in EMOTIONS:
        count = emotion_counts.get(emo, 0)
        pct = count / len(all_samples) * 100
        bar = "#" * int(pct / 2)
        print(f"  {emo:>10s}: {count:>5d} ({pct:5.1f}%) {bar}")

    # 4. Split by user
    print(f"\n[4/5] Splitting by user (seed={args.seed})...")
    train, val, test, train_u, val_u, test_u = split_by_user(
        all_samples, SPLIT_RATIO, args.seed
    )
    print(f"  Train: {len(train)} samples ({len(train_u)} users: {sorted(train_u)})")
    print(f"  Val:   {len(val)} samples ({len(val_u)} users: {sorted(val_u)})")
    print(f"  Test:  {len(test)} samples ({len(test_u)} users: {sorted(test_u)})")

    # 5. Build numpy arrays
    print(f"\n[5/5] Building numpy arrays...")
    X_train_img, X_train_lm, y_train = build_arrays(train, "Train")
    X_val_img, X_val_lm, y_val = build_arrays(val, "Val")
    X_test_img, X_test_lm, y_test = build_arrays(test, "Test")

    # Save
    print(f"\n  Saving to {output}/...")
    np.save(output / "X_train_images.npy", X_train_img)
    np.save(output / "X_train_landmarks.npy", X_train_lm)
    np.save(output / "y_train.npy", y_train)
    np.save(output / "X_val_images.npy", X_val_img)
    np.save(output / "X_val_landmarks.npy", X_val_lm)
    np.save(output / "y_val.npy", y_val)
    np.save(output / "X_test_images.npy", X_test_img)
    np.save(output / "X_test_landmarks.npy", X_test_lm)
    np.save(output / "y_test.npy", y_test)

    # Save label map
    with open(output / "label_map.json", "w") as f:
        json.dump({str(v): k for k, v in LABEL_MAP.items()}, f, indent=2)

    # Save dataset info
    info = {
        "total_samples": len(all_samples),
        "old_samples": len(old_samples),
        "new_samples": len(new_samples),
        "include_side": args.include_side,
        "split_ratio": list(SPLIT_RATIO),
        "seed": args.seed,
        "train": {"samples": len(train), "users": sorted(train_u)},
        "val": {"samples": len(val), "users": sorted(val_u)},
        "test": {"samples": len(test), "users": sorted(test_u)},
        "emotion_distribution": dict(emotion_counts),
        "image_shape": [IMG_SIZE, IMG_SIZE, 3],
        "landmark_dim": 136,
        "num_classes": len(EMOTIONS),
        "emotions": EMOTIONS,
    }
    with open(output / "dataset_info.json", "w") as f:
        json.dump(info, f, indent=2)

    # Summary
    print(f"\n{'=' * 60}")
    print(f"SELESAI!")
    print(f"  Train: {X_train_img.shape[0]} ({X_train_img.nbytes / 1e9:.2f} GB images)")
    print(f"  Val:   {X_val_img.shape[0]}")
    print(f"  Test:  {X_test_img.shape[0]}")
    print(f"  Output: {output}/")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
