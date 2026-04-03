"""
Web Tool Validasi Emosi - Expert Review
=========================================
Tool interaktif untuk ahli psikologi memvalidasi label emosi.

Usage:
    streamlit run src/tools/validation_app.py
    streamlit run src/tools/validation_app.py -- --data-dir data/validation_stratified_5pct
"""

import streamlit as st
import json
import os
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter

# ============== CONFIG ==============
EMOTIONS = ["neutral", "happy", "sad", "angry", "fearful", "disgusted", "surprised"]
EMOTION_COLORS = {
    "neutral": "#6c757d",
    "happy": "#28a745",
    "sad": "#007bff",
    "angry": "#dc3545",
    "fearful": "#9b59b6",
    "disgusted": "#8e44ad",
    "surprised": "#ffc107",
}
# ====================================


def find_data_dir():
    """Find validation data directory."""
    # Check CLI args
    if "--data-dir" in sys.argv:
        idx = sys.argv.index("--data-dir")
        if idx + 1 < len(sys.argv):
            return Path(sys.argv[idx + 1])

    # Default: look for validation dirs
    for name in ["validation_stratified_5pct", "validation_stratified_10pct",
                  "validation_full_1938", "validation"]:
        p = Path(f"data/{name}")
        if p.exists() and (p / "validation_sheet.xlsx").exists():
            return p
    return None


def load_validation_data(data_dir):
    """Load validation data from Excel."""
    import openpyxl

    xlsx_path = data_dir / "validation_sheet.xlsx"
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    samples = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        sample = {
            "no": row[0],
            "filename": row[1],
            "user_id": row[2],
            "source": row[3],
            "angle": row[4],
            "auto_label": row[5],
            "confidence": row[6],
            "scores": {
                "neutral": row[7],
                "happy": row[8],
                "sad": row[9],
                "angry": row[10],
                "fearful": row[11],
                "disgusted": row[12],
                "surprised": row[13],
            },
        }
        samples.append(sample)
    wb.close()
    return samples


def get_results_path(data_dir):
    """Path untuk simpan hasil validasi."""
    return data_dir / "expert_results.json"


def load_results(data_dir):
    """Load hasil validasi yang sudah ada."""
    path = get_results_path(data_dir)
    if path.exists():
        with open(path, "r") as f:
            return json.load(f)
    return {}


def save_results(data_dir, results):
    """Simpan hasil validasi."""
    path = get_results_path(data_dir)
    with open(path, "w") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)


def find_image(data_dir, filename):
    """Cari file gambar di folder images/."""
    images_dir = data_dir / "images"
    # Cari file yang mengandung filename
    for f in images_dir.iterdir():
        if filename in f.name:
            return f
    # Fallback: exact match
    exact = images_dir / filename
    if exact.exists():
        return exact
    return None


def main():
    st.set_page_config(
        page_title="Validasi Emosi - Expert Review",
        page_icon="🎭",
        layout="wide",
    )

    # Find data
    data_dir = find_data_dir()
    if data_dir is None:
        st.error("Data validasi tidak ditemukan. Jalankan generate_validation_set.py dulu.")
        return

    # Load data
    if "samples" not in st.session_state:
        st.session_state.samples = load_validation_data(data_dir)
        st.session_state.results = load_results(data_dir)
        st.session_state.data_dir = data_dir

    samples = st.session_state.samples
    results = st.session_state.results
    total = len(samples)
    validated = len(results)

    # ============ SIDEBAR ============
    with st.sidebar:
        st.title("Validasi Emosi")
        st.caption(f"Data: `{data_dir}`")

        # Progress
        pct = validated / total * 100 if total > 0 else 0
        st.progress(pct / 100)
        st.metric("Progress", f"{validated} / {total} ({pct:.1f}%)")

        # Stats
        if validated > 0:
            agreed = sum(1 for k, v in results.items() if v["expert_label"] == "agree")
            changed = validated - agreed
            st.metric("Setuju dengan auto-label", f"{agreed} ({agreed/validated*100:.0f}%)")
            st.metric("Dikoreksi", f"{changed} ({changed/validated*100:.0f}%)")

        st.divider()

        # Navigation
        st.subheader("Navigasi")
        nav_mode = st.radio("Mode", ["Belum divalidasi", "Semua", "Yang dikoreksi"], index=0)

        if nav_mode == "Belum divalidasi":
            filtered = [s for s in samples if str(s["no"]) not in results]
        elif nav_mode == "Yang dikoreksi":
            filtered = [s for s in samples
                        if str(s["no"]) in results and results[str(s["no"])]["expert_label"] != "agree"]
        else:
            filtered = samples

        if not filtered:
            st.success("Semua sudah divalidasi!" if nav_mode == "Belum divalidasi"
                       else "Tidak ada data.")
            # Show summary
            if validated > 0:
                st.divider()
                st.subheader("Ringkasan")
                corrections = Counter()
                for k, v in results.items():
                    if v["expert_label"] != "agree":
                        corrections[v["expert_label"]] += 1
                if corrections:
                    st.write("Koreksi per emosi:")
                    for emo, count in corrections.most_common():
                        st.write(f"  - {emo}: {count}")
            return

        # Filter by emotion
        st.divider()
        filter_emo = st.selectbox("Filter emosi", ["Semua"] + EMOTIONS, index=0)
        if filter_emo != "Semua":
            filtered = [s for s in filtered if s["auto_label"] == filter_emo]

        if not filtered:
            st.info("Tidak ada data untuk filter ini.")
            return

        # Sample selector
        sample_idx = st.number_input(
            f"Sample ({len(filtered)} tersisa)",
            min_value=0, max_value=len(filtered) - 1, value=0, step=1
        )

    # ============ MAIN AREA ============
    sample = filtered[sample_idx]
    sample_key = str(sample["no"])
    existing = results.get(sample_key, {})

    # Header
    col_title, col_nav = st.columns([3, 1])
    with col_title:
        st.title(f"Sample #{sample['no']}")
        st.caption(f"User: {sample['user_id']} | Source: {sample['source']} | Angle: {sample['angle']}")

    with col_nav:
        st.write("")
        col_prev, col_next = st.columns(2)
        with col_prev:
            if st.button("⬅ Prev", use_container_width=True, disabled=sample_idx == 0):
                st.session_state.nav_idx = sample_idx - 1
                st.rerun()
        with col_next:
            if st.button("Next ➡", use_container_width=True, disabled=sample_idx >= len(filtered) - 1):
                st.session_state.nav_idx = sample_idx + 1
                st.rerun()

    st.divider()

    # Image + Label side by side
    col_img, col_info = st.columns([1, 1])

    with col_img:
        img_path = find_image(data_dir, sample["filename"])
        if img_path and img_path.exists():
            st.image(str(img_path), caption=sample["filename"], width=400)
        else:
            st.warning(f"Gambar tidak ditemukan: {sample['filename']}")

    with col_info:
        # Auto label
        auto_label = sample["auto_label"]
        confidence = sample["confidence"]
        color = EMOTION_COLORS.get(auto_label, "#333")

        st.markdown(f"### Label Otomatis: "
                    f"<span style='color:{color}; font-size:1.5em; font-weight:bold'>"
                    f"{auto_label.upper()}</span>", unsafe_allow_html=True)
        st.metric("Confidence", f"{confidence:.4f}" if confidence else "N/A")

        # Score bars
        st.write("**Skor per emosi:**")
        scores = sample["scores"]
        for emo in EMOTIONS:
            score = scores.get(emo, 0)
            if score is None:
                score = 0
            score = float(score)
            bar_color = EMOTION_COLORS.get(emo, "#333")
            is_max = (emo == auto_label)
            label = f"**{emo}**" if is_max else emo
            st.progress(min(score, 1.0), text=f"{label}: {score:.6f}")

    st.divider()

    # ============ EXPERT INPUT ============
    st.subheader("Validasi Ahli")

    # Quick buttons
    st.write("**Apakah label otomatis sudah benar?**")

    col_agree, *col_emotions = st.columns(len(EMOTIONS) + 1)

    with col_agree:
        if st.button("✅ Setuju", use_container_width=True, type="primary"):
            results[sample_key] = {
                "expert_label": "agree",
                "auto_label": auto_label,
                "notes": "",
                "timestamp": datetime.now().isoformat(),
            }
            save_results(data_dir, results)
            st.session_state.results = results
            st.rerun()

    for col, emo in zip(col_emotions, EMOTIONS):
        with col:
            btn_type = "secondary"
            if emo == auto_label:
                continue  # Skip auto label (already covered by "Setuju")
            if st.button(emo.capitalize(), use_container_width=True, type=btn_type):
                results[sample_key] = {
                    "expert_label": emo,
                    "auto_label": auto_label,
                    "notes": "",
                    "timestamp": datetime.now().isoformat(),
                }
                save_results(data_dir, results)
                st.session_state.results = results
                st.rerun()

    # Notes
    notes = st.text_input("Catatan (opsional)",
                          value=existing.get("notes", ""),
                          key=f"notes_{sample_key}")
    if notes and sample_key in results:
        results[sample_key]["notes"] = notes
        save_results(data_dir, results)

    # Show current status
    if sample_key in results:
        r = results[sample_key]
        if r["expert_label"] == "agree":
            st.success(f"✅ Sudah divalidasi: Setuju dengan '{auto_label}'")
        else:
            st.warning(f"✏️ Dikoreksi: '{auto_label}' → '{r['expert_label']}'")

    # ============ FOOTER ============
    st.divider()
    col_export, col_reset = st.columns([3, 1])
    with col_export:
        if validated > 0 and st.button("📊 Export Hasil ke CSV"):
            export_path = data_dir / "expert_results.csv"
            import csv
            with open(export_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["no", "filename", "user_id", "auto_label",
                                 "confidence", "expert_label", "notes", "timestamp"])
                for s in samples:
                    key = str(s["no"])
                    if key in results:
                        r = results[key]
                        final_label = s["auto_label"] if r["expert_label"] == "agree" else r["expert_label"]
                        writer.writerow([
                            s["no"], s["filename"], s["user_id"], s["auto_label"],
                            s["confidence"], final_label, r.get("notes", ""), r.get("timestamp", "")
                        ])
            st.success(f"Exported ke {export_path}")

    with col_reset:
        if sample_key in results:
            if st.button("🗑️ Reset ini", type="secondary"):
                del results[sample_key]
                save_results(data_dir, results)
                st.session_state.results = results
                st.rerun()


if __name__ == "__main__":
    main()
